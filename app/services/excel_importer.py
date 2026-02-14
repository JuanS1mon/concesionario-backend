"""
Servicio de importación de datos de mercado desde archivos Excel (.xlsx).
Permite cargar datos manualmente cuando el scraping no está disponible.

Soporta dos formatos de hoja:
  - "Datos de Mercado": listings individuales (como MercadoLibre/Kavak/deRuedas)
  - "Precios de Referencia": rangos de precios por marca/modelo/año/versión
"""
import logging
import re
from datetime import datetime
from typing import Optional, BinaryIO
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models.pricing import MarketRawListing

logger = logging.getLogger(__name__)

# Mapeo de nombres de columna esperados → campo interno
COLUMN_MAP_MERCADO = {
    "marca": "marca_raw",
    "modelo": "modelo_raw",
    "año": "anio",
    "anio": "anio",
    "ano": "anio",
    "kilometraje": "km",
    "km": "km",
    "kilómetros": "km",
    "kilometros": "km",
    "precio": "precio",
    "moneda": "moneda",
    "ubicación": "ubicacion",
    "ubicacion": "ubicacion",
    "fuente": "fuente_custom",
    "url": "url",
    "combustible": "combustible",
    "observaciones": "observaciones",
    "titulo": "titulo",
    "título": "titulo",
}

COLUMN_MAP_REFERENCIA = {
    "marca": "marca_raw",
    "modelo": "modelo_raw",
    "año": "anio",
    "anio": "anio",
    "ano": "anio",
    "versión": "version",
    "version": "version",
    "precio mínimo": "precio_min",
    "precio minimo": "precio_min",
    "precio_min": "precio_min",
    "precio máximo": "precio_max",
    "precio maximo": "precio_max",
    "precio_max": "precio_max",
    "precio": "precio",
    "moneda": "moneda",
    "observaciones": "observaciones",
}


def _normalize_header(header: str) -> str:
    """Normaliza un nombre de columna para matching flexible."""
    if not header:
        return ""
    return header.strip().lower().replace("_", " ")


def _parse_number(value) -> Optional[float]:
    """Convierte un valor a número, limpiando formatos comunes."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        clean = value.strip().replace("$", "").replace(".", "").replace(",", "").strip()
        clean = re.sub(r"[^\d.-]", "", clean)
        try:
            return float(clean) if clean else None
        except ValueError:
            return None
    return None


def _parse_int(value) -> Optional[int]:
    """Convierte un valor a entero."""
    num = _parse_number(value)
    return int(num) if num is not None else None


def _detect_sheet_type(headers: list[str]) -> str:
    """Detecta si una hoja es de 'mercado' o 'referencia' por sus columnas."""
    headers_lower = [_normalize_header(h) for h in headers if h]
    if any("precio mínimo" in h or "precio minimo" in h or "precio_min" in h for h in headers_lower):
        return "referencia"
    if any("km" in h or "kilometraje" in h or "kilómetros" in h for h in headers_lower):
        return "mercado"
    # Default: si tiene marca y precio, asumir mercado
    return "mercado"


def _map_headers(headers: list[str], column_map: dict) -> dict[int, str]:
    """
    Mapea índices de columna a nombres de campo internos.
    Retorna: {col_index: field_name}
    """
    mapping = {}
    for idx, header in enumerate(headers):
        norm = _normalize_header(header)
        if norm in column_map:
            mapping[idx] = column_map[norm]
    return mapping


def _process_mercado_row(row_data: dict, row_num: int) -> Optional[dict]:
    """Procesa una fila de la hoja 'Datos de Mercado'."""
    marca = row_data.get("marca_raw", "")
    modelo = row_data.get("modelo_raw", "")
    precio = _parse_number(row_data.get("precio"))
    anio = _parse_int(row_data.get("anio"))

    if not marca or not precio:
        logger.debug(f"Fila {row_num}: sin marca o precio, saltando")
        return None

    km = _parse_int(row_data.get("km"))
    moneda = str(row_data.get("moneda", "ARS")).strip().upper()
    if moneda not in ("ARS", "USD"):
        moneda = "ARS"

    ubicacion = str(row_data.get("ubicacion", "")).strip() if row_data.get("ubicacion") else ""
    fuente_custom = str(row_data.get("fuente_custom", "")).strip() if row_data.get("fuente_custom") else "excel_manual"
    url = str(row_data.get("url", "")).strip() if row_data.get("url") else None
    observaciones = str(row_data.get("observaciones", "")).strip() if row_data.get("observaciones") else ""
    combustible = str(row_data.get("combustible", "")).strip() if row_data.get("combustible") else ""

    # Construir título
    titulo_parts = [str(marca), str(modelo)]
    if anio:
        titulo_parts.append(str(anio))
    if combustible:
        titulo_parts.append(combustible)
    if observaciones:
        titulo_parts.append(f"- {observaciones}")

    titulo = row_data.get("titulo") or " ".join(titulo_parts)

    # URL única para deduplicación
    if not url:
        url = f"excel://{marca}/{modelo}/{anio or 'sin-anio'}/{row_num}"

    return {
        "fuente": f"excel:{fuente_custom}" if fuente_custom != "excel_manual" else "excel",
        "url": url,
        "titulo": str(titulo).strip(),
        "marca_raw": str(marca).strip(),
        "modelo_raw": str(modelo).strip(),
        "anio": anio,
        "km": km,
        "precio": precio,
        "moneda": moneda,
        "ubicacion": ubicacion,
        "imagen_url": "",
        "activo": True,
        "procesado": False,
        "fecha_scraping": datetime.utcnow(),
    }


def _process_referencia_row(row_data: dict, row_num: int) -> Optional[dict]:
    """Procesa una fila de la hoja 'Precios de Referencia'."""
    marca = row_data.get("marca_raw", "")
    modelo = row_data.get("modelo_raw", "")
    anio = _parse_int(row_data.get("anio"))

    precio_min = _parse_number(row_data.get("precio_min"))
    precio_max = _parse_number(row_data.get("precio_max"))
    precio_unico = _parse_number(row_data.get("precio"))

    if not marca:
        return None

    # Calcular precio: promedio de min/max, o el precio único
    if precio_min and precio_max:
        precio = (precio_min + precio_max) / 2
    elif precio_unico:
        precio = precio_unico
    elif precio_min:
        precio = precio_min
    elif precio_max:
        precio = precio_max
    else:
        logger.debug(f"Fila {row_num}: sin precios, saltando")
        return None

    moneda = str(row_data.get("moneda", "ARS")).strip().upper()
    if moneda not in ("ARS", "USD"):
        moneda = "ARS"

    version = str(row_data.get("version", "")).strip() if row_data.get("version") else ""
    observaciones = str(row_data.get("observaciones", "")).strip() if row_data.get("observaciones") else ""

    titulo_parts = [str(marca), str(modelo)]
    if anio:
        titulo_parts.append(str(anio))
    if version:
        titulo_parts.append(version)
    if precio_min and precio_max:
        titulo_parts.append(f"(${precio_min:,.0f} - ${precio_max:,.0f})")

    url = f"excel-ref://{marca}/{modelo}/{anio or 'sin-anio'}/{version or 'base'}/{row_num}"

    return {
        "fuente": "excel_referencia",
        "url": url,
        "titulo": " ".join(titulo_parts),
        "marca_raw": str(marca).strip(),
        "modelo_raw": str(modelo).strip() + (f" {version}" if version else ""),
        "anio": anio,
        "km": None,
        "precio": precio,
        "moneda": moneda,
        "ubicacion": "Nacional",
        "imagen_url": "",
        "activo": True,
        "procesado": False,
        "fecha_scraping": datetime.utcnow(),
    }


def importar_excel(
    db: Session,
    file_content: bytes,
    filename: str = "upload.xlsx",
    sobrescribir: bool = False,
) -> dict:
    """
    Importa datos de mercado desde un archivo Excel.

    Args:
        db: Sesión de base de datos
        file_content: Contenido binario del archivo Excel
        filename: Nombre del archivo (para logging)
        sobrescribir: Si True, elimina datos previos de fuente 'excel*' antes de importar

    Returns:
        dict con stats: {importados, duplicados, errores, hojas_procesadas, detalles}
    """
    stats = {
        "importados": 0,
        "duplicados": 0,
        "errores": 0,
        "filas_sin_datos": 0,
        "hojas_procesadas": [],
        "detalles": [],
    }

    try:
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Error abriendo Excel '{filename}': {e}")
        stats["errores"] += 1
        stats["detalles"].append(f"Error al abrir el archivo: {str(e)}")
        return stats

    # Si sobrescribir, eliminar datos previos de importación Excel
    if sobrescribir:
        deleted = db.query(MarketRawListing).filter(
            MarketRawListing.fuente.like("excel%")
        ).delete(synchronize_session=False)
        db.commit()
        logger.info(f"Eliminados {deleted} registros previos de importación Excel")
        stats["detalles"].append(f"Eliminados {deleted} registros previos")

    # Cargar URLs existentes para deduplicación
    existing_urls = set(
        u for (u,) in db.query(MarketRawListing.url).filter(
            MarketRawListing.fuente.like("excel%")
        ).all()
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"Procesando hoja: '{sheet_name}'")

        # Saltar hoja de instrucciones
        if "instruc" in sheet_name.lower():
            logger.info(f"Saltando hoja de instrucciones: '{sheet_name}'")
            continue

        # Leer headers (primera fila)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            logger.warning(f"Hoja '{sheet_name}' vacía")
            continue

        headers = [str(h).strip() if h else "" for h in header_row]
        sheet_type = _detect_sheet_type(headers)

        if sheet_type == "referencia":
            column_map = COLUMN_MAP_REFERENCIA
            process_fn = _process_referencia_row
        else:
            column_map = COLUMN_MAP_MERCADO
            process_fn = _process_mercado_row

        header_mapping = _map_headers(headers, column_map)

        if not header_mapping:
            logger.warning(f"Hoja '{sheet_name}': no se pudieron mapear columnas")
            stats["detalles"].append(f"Hoja '{sheet_name}': columnas no reconocidas")
            continue

        logger.info(f"Hoja '{sheet_name}': tipo={sheet_type}, columnas mapeadas={list(header_mapping.values())}")

        sheet_stats = {"importados": 0, "duplicados": 0, "errores": 0}
        nuevos = []

        for row_num, row in enumerate(rows_iter, start=2):
            try:
                # Construir dict de la fila
                row_data = {}
                for col_idx, field_name in header_mapping.items():
                    if col_idx < len(row):
                        row_data[field_name] = row[col_idx]

                # Procesar fila
                result = process_fn(row_data, row_num)
                if result is None:
                    stats["filas_sin_datos"] += 1
                    continue

                # Deduplicar
                if result["url"] in existing_urls:
                    sheet_stats["duplicados"] += 1
                    continue

                nuevos.append(MarketRawListing(**result))
                existing_urls.add(result["url"])
                sheet_stats["importados"] += 1

            except Exception as e:
                logger.error(f"Hoja '{sheet_name}', fila {row_num}: {e}")
                sheet_stats["errores"] += 1

        # Guardar en lotes
        if nuevos:
            BATCH_SIZE = 50
            for i in range(0, len(nuevos), BATCH_SIZE):
                batch = nuevos[i:i + BATCH_SIZE]
                db.add_all(batch)
                db.commit()

        for k in ("importados", "duplicados", "errores"):
            stats[k] += sheet_stats[k]

        stats["hojas_procesadas"].append({
            "nombre": sheet_name,
            "tipo": sheet_type,
            "importados": sheet_stats["importados"],
            "duplicados": sheet_stats["duplicados"],
            "errores": sheet_stats["errores"],
        })

        logger.info(f"Hoja '{sheet_name}': {sheet_stats}")

    wb.close()
    logger.info(f"Importación Excel completada: {stats}")
    return stats
