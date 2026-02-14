#!/usr/bin/env python
"""
Genera un archivo Excel de ejemplo con datos de mercado automotor.
Este archivo sirve como plantilla para importar datos manualmente
cuando las fuentes de scraping no están disponibles.

Uso:
    python generate_sample_excel.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sample_excel(output_path: str = "datos_mercado_ejemplo.xlsx"):
    wb = Workbook()

    # ─── Hoja 1: Datos de Mercado (listings individuales) ─────────
    ws1 = wb.active
    ws1.title = "Datos de Mercado"

    headers_mercado = [
        "Marca", "Modelo", "Año", "Kilometraje", "Precio",
        "Moneda", "Ubicación", "Fuente", "URL", "Combustible", "Observaciones",
    ]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escribir headers
    for col, header in enumerate(headers_mercado, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Datos de ejemplo (variados y realistas para Argentina)
    sample_data = [
        ("Toyota", "Corolla", 2022, 35000, 28500000, "ARS", "CABA", "manual", "", "Nafta", "Único dueño"),
        ("Toyota", "Corolla", 2021, 52000, 25800000, "ARS", "Córdoba", "manual", "", "Nafta", ""),
        ("Toyota", "Corolla", 2023, 18000, 32000000, "ARS", "Mendoza", "manual", "", "Nafta", "Con service oficial"),
        ("Toyota", "Hilux", 2022, 45000, 42000000, "ARS", "Tucumán", "manual", "", "Diesel", "4x4 SRX"),
        ("Toyota", "Hilux", 2021, 68000, 38500000, "ARS", "CABA", "manual", "", "Diesel", "4x2 SR"),
        ("Toyota", "Etios", 2019, 72000, 15200000, "ARS", "Buenos Aires", "manual", "", "Nafta", ""),
        ("Ford", "Ranger", 2022, 40000, 39800000, "ARS", "Santa Fe", "manual", "", "Diesel", "XLT 4x4"),
        ("Ford", "Ranger", 2023, 22000, 45000000, "ARS", "CABA", "manual", "", "Diesel", "Limited"),
        ("Ford", "Focus", 2018, 95000, 16500000, "ARS", "Rosario", "manual", "", "Nafta", "SE 2.0"),
        ("Ford", "Territory", 2023, 15000, 35000000, "ARS", "CABA", "manual", "", "Nafta", "Titanium"),
        ("Volkswagen", "Amarok", 2022, 38000, 44000000, "ARS", "Mendoza", "manual", "", "Diesel", "V6 Extreme"),
        ("Volkswagen", "Polo", 2021, 45000, 18500000, "ARS", "Córdoba", "manual", "", "Nafta", "Trendline"),
        ("Volkswagen", "T-Cross", 2023, 12000, 29500000, "ARS", "CABA", "manual", "", "Nafta", "Highline"),
        ("Chevrolet", "Cruze", 2022, 30000, 24000000, "ARS", "Buenos Aires", "manual", "", "Nafta", "LTZ AT"),
        ("Chevrolet", "Tracker", 2023, 20000, 31000000, "ARS", "CABA", "manual", "", "Nafta", "Premier"),
        ("Chevrolet", "S10", 2021, 55000, 37500000, "ARS", "Entre Ríos", "manual", "", "Diesel", "High Country"),
        ("Fiat", "Cronos", 2022, 40000, 17800000, "ARS", "Córdoba", "manual", "", "Nafta", "Precision CVT"),
        ("Fiat", "Pulse", 2023, 18000, 23500000, "ARS", "CABA", "manual", "", "Nafta", "Impetus"),
        ("Renault", "Duster", 2021, 50000, 21000000, "ARS", "Mendoza", "manual", "", "Nafta", "Outsider 4x4"),
        ("Renault", "Sandero", 2020, 62000, 14500000, "ARS", "Tucumán", "manual", "", "Nafta", "Stepway"),
        ("Peugeot", "208", 2023, 15000, 24000000, "ARS", "CABA", "manual", "", "Nafta", "Feline AT"),
        ("Peugeot", "2008", 2022, 32000, 27500000, "ARS", "Rosario", "manual", "", "Nafta", "Allure"),
        ("Honda", "HR-V", 2022, 28000, 33000000, "ARS", "CABA", "manual", "", "Nafta", "EXL CVT"),
        ("Honda", "Civic", 2023, 10000, 38000000, "ARS", "Buenos Aires", "manual", "", "Nafta", "EXL Turbo"),
        ("Hyundai", "Tucson", 2022, 35000, 34500000, "ARS", "CABA", "manual", "", "Nafta", "GL 4x2"),
        ("Nissan", "Kicks", 2021, 42000, 22000000, "ARS", "Mendoza", "manual", "", "Nafta", "Exclusive CVT"),
        ("Jeep", "Renegade", 2022, 25000, 29000000, "ARS", "Córdoba", "manual", "", "Nafta", "Sport AT"),
        ("Jeep", "Compass", 2023, 18000, 42000000, "ARS", "CABA", "manual", "", "Nafta", "Limited TD"),
        # Ejemplo en USD
        ("Toyota", "SW4", 2023, 20000, 52000, "USD", "CABA", "manual", "", "Diesel", "SRX 4x4 - Precio en USD"),
        ("Mercedes-Benz", "Clase A", 2022, 18000, 45000, "USD", "CABA", "manual", "", "Nafta", "A200 Progressive"),
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 5:  # Precio
                cell.number_format = '#,##0'
            if col_idx == 4:  # KM
                cell.number_format = '#,##0'

    # Ajustar anchos de columna
    col_widths = [15, 15, 8, 14, 16, 10, 16, 10, 40, 12, 30]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # ─── Hoja 2: Precios de Referencia ────────────────────────────
    ws2 = wb.create_sheet("Precios de Referencia")

    headers_ref = [
        "Marca", "Modelo", "Año", "Versión",
        "Precio Mínimo", "Precio Máximo", "Moneda", "Observaciones",
    ]

    ref_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")

    for col, header in enumerate(headers_ref, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = ref_fill
        cell.alignment = header_align
        cell.border = thin_border

    ref_data = [
        ("Toyota", "Corolla", 2023, "XLi 2.0 CVT", 23064600, 26500000, "ARS", ""),
        ("Toyota", "Corolla", 2023, "SEG 2.0 CVT", 27000000, 31558100, "ARS", ""),
        ("Toyota", "Corolla", 2022, "XLi 2.0 CVT", 21911400, 24500000, "ARS", ""),
        ("Toyota", "Corolla", 2022, "SEG 2.0 CVT", 25000000, 28481200, "ARS", ""),
        ("Toyota", "Hilux", 2023, "SR 4x2 2.8 TDI", 35000000, 38000000, "ARS", ""),
        ("Toyota", "Hilux", 2023, "SRX 4x4 2.8 TDI AT", 42000000, 48000000, "ARS", ""),
        ("Ford", "Ranger", 2023, "XLT 4x4 3.0 TDI", 38000000, 42000000, "ARS", ""),
        ("Ford", "Ranger", 2023, "Limited 4x4 3.0 TDI", 44000000, 50000000, "ARS", ""),
        ("Volkswagen", "Amarok", 2023, "Comfortline 4x2", 36000000, 40000000, "ARS", ""),
        ("Volkswagen", "Amarok", 2023, "V6 Extreme 4x4", 50000000, 56000000, "ARS", ""),
        ("Chevrolet", "Cruze", 2023, "LTZ AT", 28000000, 32000000, "ARS", ""),
        ("Fiat", "Cronos", 2023, "Precision 1.3 CVT", 18000000, 21000000, "ARS", ""),
        ("Peugeot", "208", 2023, "Feline 1.6 AT", 22000000, 26000000, "ARS", ""),
        ("Honda", "HR-V", 2023, "EXL 1.5T CVT", 34000000, 38000000, "ARS", ""),
        ("Renault", "Duster", 2023, "Outsider 1.3T 4x4", 25000000, 29000000, "ARS", ""),
    ]

    for row_idx, row_data in enumerate(ref_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (5, 6):  # Precios
                cell.number_format = '#,##0'

    ref_widths = [15, 15, 8, 25, 16, 16, 10, 30]
    for i, width in enumerate(ref_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # ─── Hoja 3: Instrucciones ────────────────────────────────────
    ws3 = wb.create_sheet("Instrucciones")

    instructions = [
        ("INSTRUCCIONES DE USO", ""),
        ("", ""),
        ("Hoja 'Datos de Mercado':", ""),
        ("  Marca", "Nombre de la marca del vehículo (Toyota, Ford, Volkswagen, etc.)"),
        ("  Modelo", "Nombre del modelo (Corolla, Ranger, Amarok, etc.)"),
        ("  Año", "Año de fabricación (número entero, ej: 2023)"),
        ("  Kilometraje", "Kilómetros recorridos (dejar vacío si no aplica)"),
        ("  Precio", "Precio numérico SIN puntos ni signos (ej: 28500000)"),
        ("  Moneda", "ARS para pesos argentinos, USD para dólares"),
        ("  Ubicación", "Ciudad o provincia donde se encuentra el vehículo"),
        ("  Fuente", "Origen del dato: manual, concesionario, portal, revista, etc."),
        ("  URL", "Link al aviso original (opcional)"),
        ("  Combustible", "Nafta, Diesel, GNC, Eléctrico, Híbrido (opcional)"),
        ("  Observaciones", "Cualquier detalle adicional (opcional)"),
        ("", ""),
        ("Hoja 'Precios de Referencia':", ""),
        ("  Marca/Modelo/Año/Versión", "Identifican el vehículo específico"),
        ("  Precio Mínimo", "Valor más bajo observado en el mercado"),
        ("  Precio Máximo", "Valor más alto observado en el mercado"),
        ("  Moneda", "ARS o USD"),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("  1.", "Puede usar cualquiera de las dos hojas o ambas"),
        ("  2.", "Los campos Marca, Modelo, Año y Precio son OBLIGATORIOS"),
        ("  3.", "Si la Moneda está vacía, se asume ARS (pesos argentinos)"),
        ("  4.", "Los datos se procesan igual que el scraping automático"),
        ("  5.", "Se puede subir este archivo desde Admin > Pricing > Importar Excel"),
    ]

    title_font = Font(bold=True, size=14, color="2E86AB")
    subtitle_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        cell_a = ws3.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws3.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = title_font
        elif col_a.endswith(":") and not col_a.startswith(" "):
            cell_a.font = subtitle_font
        elif col_a.startswith("  "):
            cell_a.font = Font(bold=True, size=10)
            cell_b.font = normal_font
        else:
            cell_a.font = normal_font

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 60

    # Guardar
    wb.save(output_path)
    print(f"Archivo Excel de ejemplo creado: {output_path}")
    print(f"  - {len(sample_data)} registros en 'Datos de Mercado'")
    print(f"  - {len(ref_data)} registros en 'Precios de Referencia'")
    print(f"  - Hoja 'Instrucciones' con guía de uso")
    return output_path


if __name__ == "__main__":
    create_sample_excel()
